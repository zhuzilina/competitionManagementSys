from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from rest_framework import viewsets
from datetime import datetime
from django.db.models import Prefetch
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Award
from .serializers import AwardSerializer
from .serializers import AwardReportSerializer
from userManage.permissions import IsCompAdminOrReadOnly,IsCompAdmin
from django.db.models import Count,Q,F
from django.db.models.functions import ExtractYear

User = get_user_model()

class AwardViewSet(viewsets.ModelViewSet):
    queryset = Award.objects.all().select_related('competition', 'certificate').prefetch_related('participants', 'instructors')
    serializer_class = AwardSerializer
    permission_classes = [IsCompAdminOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user_query_id = self.request.query_params.get('user_id')

        # 如果传了 user_id='me'，自动替换为当前登录用户 ID
        if user_query_id == 'me':
            if self.request.user.is_authenticated:
                user_query_id = self.request.user.user_id  # 假设你的字段名是这个
            else:
                return Award.objects.none()

        if user_query_id:
            queryset = queryset.filter(
                Q(participants__user_id=user_query_id) |
                Q(instructors__user_id=user_query_id)
            ).distinct()

        return queryset

    def perform_create(self, serializer):
        # 自动关联当前登录用户为录入人
        serializer.save(creator=self.request.user)

    def perform_destroy(self, instance):
        cert = instance.certificate
        instance.delete()  # 删除获奖记录
        if cert:
            cert.delete()


class AwardReportView(APIView):
    """
    获奖统计报表
    网页查看
    GET /award/report/?group_by=student&start_date=2025-01-01
    下载报表
    GET /award/report/?group_by=student&start_date=2025-01-01&format=excel
    """
    permission_classes = [IsCompAdmin]

    def get(self, request):
        group_by = request.query_params.get('group_by', 'student')
        # 获取日期范围参数
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # 1. 构建奖项的过滤基础查询集
        award_queryset = Award.objects.select_related('competition')

        if start_date:
            award_queryset = award_queryset.filter(award_date__gte=start_date)
        if end_date:
            award_queryset = award_queryset.filter(award_date__lte=end_date)

        report_data = []

        if group_by == 'student':
            # 使用 Prefetch 对象，将过滤后的奖项存入 'filtered_awards' 属性中
            users = User.objects.filter(
                student_awards__in=award_queryset
            ).distinct().prefetch_related(
                'profile',
                Prefetch('student_awards', queryset=award_queryset, to_attr='filtered_awards')
            )

            for user in users:
                # 注意这里使用 to_attr 指定的 'filtered_awards'
                report_data.append(self._format_user_data(user, user.filtered_awards))

        elif group_by == 'teacher':
            users = User.objects.filter(
                teacher_awards__in=award_queryset
            ).distinct().prefetch_related(
                'profile',
                Prefetch('teacher_awards', queryset=award_queryset, to_attr='filtered_awards')
            )

            for user in users:
                report_data.append(self._format_user_data(user, user.filtered_awards))

        serializer = AwardReportSerializer(report_data, many=True)
        return Response(serializer.data)

    def _generate_excel(self, data, group_by):
        """生成并返回 Excel 文件流"""
        wb = Workbook()
        ws = wb.active
        ws.title = "获奖报表"

        # 表头
        headers = ['学号/工号', '姓名', '院系', '专业/班级/职称', '获奖明细']
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充数据
        for item in data:
            # 将多条获奖信息合并为一个字符串
            awards_text = "\n".join([
                f"[{a.award_date}] {a.competition.title} - {a.award_level}"
                for a in item['awards']
            ])

            row = [
                item['user_id'],
                item['real_name'],
                item['department'],
                f"{item['major']}/{item['clazz']}/{item['title']}",
                awards_text
            ]
            ws.append(row)

            # 设置单元格换行（用于显示多条获奖）
            ws.cell(row=ws.max_row, column=5).alignment = Alignment(wrapText=True)

        # 设置列宽
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['E'].width = 60

        # 构建响应
        filename = f"award_report_{group_by}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _format_user_data(self, user, awards):
        # 保持不变，但传入的是过滤后的 awards 列表
        profile = getattr(user, 'profile', None)
        return {
            "user_id": user.user_id,
            "real_name": profile.real_name if profile else "未填写",
            "department": profile.department if profile else "-",
            "major": getattr(profile, 'major', '-'),
            "clazz": getattr(profile, 'clazz', '-'),
            "title": getattr(profile, 'title', '-'),
            "awards": awards
        }


class AwardStatisticsView(APIView):
    """
    获奖信息多维度统计API
    """
    permission_classes = [IsCompAdmin]

    def get(self, request):
        # 1. 基础总数统计
        total_awards = Award.objects.count()

        # 2. 按竞赛类别统计 (基于 CompetitionCategory)
        category_stats = Award.objects.values(
            name=F('competition__category__name')
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # 3. 按竞赛级别统计 (基于 CompetitionLevel)
        level_stats = Award.objects.values(
            name=F('competition__level__name')
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # 4. 按年度统计及环比计算
        # 提取日期中的年份进行分组
        yearly_stats_query = Award.objects.annotate(
            year=ExtractYear('award_date')
        ).values('year').annotate(
            count=Count('id')
        ).order_by('year')

        yearly_data = list(yearly_stats_query)
        # 计算变动率 (Growth Rate)
        for i in range(len(yearly_data)):
            if i > 0 and yearly_data[i - 1]['count'] > 0:
                prev = yearly_data[i - 1]['count']
                curr = yearly_data[i]['count']
                growth_rate = ((curr - prev) / prev) * 100
                yearly_data[i]['growth_rate'] = f"{round(growth_rate, 2)}%"
            else:
                yearly_data[i]['growth_rate'] = "0%"

        # 5. 按学院部门统计
        # 注意：这里需要跨两层：Award -> participants (User) -> profile (Profile)
        dept_stats = Award.objects.values(
            name=F('participants__profile__department')
        ).annotate(
            count=Count('id', distinct=True)  # 使用distinct防止一个奖项多个学生导致重复计算
        ).exclude(name__isnull=True).order_by('-count')

        # 6. 人员总数统计 (去重)
        total_students = Award.objects.aggregate(
            count=Count('participants', distinct=True)
        )['count']

        total_instructors = Award.objects.aggregate(
            count=Count('instructors', distinct=True)
        )['count']

        # 7. 获奖等级分布 (金奖、一等奖等)
        level_distribution = Award.objects.values('award_level').annotate(
            count=Count('id')
        ).order_by('-count')

        # 封装结果
        data = {
            "summary": {
                "total_awards": total_awards,
                "total_students": total_students,
                "total_instructors": total_instructors,
            },
            "by_category": category_stats,
            "by_level": level_stats,
            "by_department": dept_stats,
            "by_year": yearly_data,
            "by_award_rank": level_distribution
        }

        return Response(data)